import os
import re
import openpyxl
from src.services.db_service import get_connection

def importar_ficha_excel(filepath: str, bloque_nombre: str = ""):
    """
    Lee un archivo Excel con la ficha individual de un departamento y
    guarda en la BD dejando únicamente la numeración del Block.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active

        filename = os.path.basename(filepath)
        depto_val = filename.split()[0].strip() if filename else ""

        # Extraer únicamente los dígitos del bloque (ej: "BLOKC 2613" -> "2613")
        texto_bloque = bloque_nombre or str(sheet.cell(row=3, column=1).value or "")
        numeros_bloque = re.findall(r'\d+', texto_bloque)
        bloque_val = numeros_bloque[0] if numeros_bloque else texto_bloque.strip()

        fojas = str(sheet.cell(row=3, column=7).value or "").strip()
        numero_insc = str(sheet.cell(row=3, column=8).value or "").strip()
        ano_insc = sheet.cell(row=3, column=9).value
        try:
            ano_insc = int(ano_insc) if ano_insc else None
        except (ValueError, TypeError):
            ano_insc = None

        rol_sii = str(sheet.cell(row=3, column=10).value or "").strip()
        avaluo = sheet.cell(row=3, column=11).value
        try:
            avaluo = float(avaluo) if avaluo else None
        except (ValueError, TypeError):
            avaluo = None

        obs_val = str(sheet.cell(row=21, column=2).value or sheet.cell(row=20, column=2).value or "").strip()

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO departamentos (bloque, numero_depto, fojas, numero_inscripcion, ano_inscripcion, rol_sii, avaluo_fiscal, observaciones)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(bloque, numero_depto) DO UPDATE SET
                fojas = excluded.fojas,
                numero_inscripcion = excluded.numero_inscripcion,
                ano_inscripcion = excluded.ano_inscripcion,
                rol_sii = excluded.rol_sii,
                avaluo_fiscal = excluded.avaluo_fiscal,
                observaciones = excluded.observaciones;
        """, (bloque_val, depto_val, fojas, numero_insc, ano_insc, rol_sii, avaluo, obs_val))

        cursor.execute("SELECT id FROM departamentos WHERE bloque = ? AND numero_depto = ?;", (bloque_val, depto_val))
        depto_id = cursor.fetchone()["id"]

        cursor.execute("DELETE FROM integrantes WHERE departamento_id = ?;", (depto_id,))

        for row_idx in range(6, 21):
            parentesco = str(sheet.cell(row=row_idx, column=1).value or "").strip()
            nombres = str(sheet.cell(row=row_idx, column=2).value or "").strip()
            ap_paterno = str(sheet.cell(row=row_idx, column=3).value or "").strip()
            ap_materno = str(sheet.cell(row=row_idx, column=4).value or "").strip()
            rut = str(sheet.cell(row=row_idx, column=6).value or "").strip()
            asistencia = str(sheet.cell(row=row_idx, column=14).value or sheet.cell(row=row_idx, column=13).value or "NO").strip()

            if parentesco and parentesco not in ["OBCERBACIONES", "OBSERVACIONES", "None"]:
                if nombres or rut or ap_paterno:
                    cursor.execute("""
                        INSERT INTO integrantes (departamento_id, parentesco, nombres, apellido_paterno, apellido_materno, rut, asistencia_reuniones)
                        VALUES (?, ?, ?, ?, ?, ?, ?);
                    """, (depto_id, parentesco, nombres, ap_paterno, ap_materno, rut, asistencia))

        conn.commit()
        conn.close()
        return True, f"Ficha {depto_val} (Block {bloque_val}) importada con éxito."

    except Exception as e:
        return False, f"Error al procesar {filepath}: {str(e)}"

def importar_carpeta_bloque(folder_path: str):
    bloque_nombre = os.path.basename(folder_path)
    archivos = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    exitos = 0
    errores = 0

    for archivo in archivos:
        full_path = os.path.join(folder_path, archivo)
        ok, msg = importar_ficha_excel(full_path, bloque_nombre)
        if ok:
            exitos += 1
        else:
            errores += 1

    return exitos, errores, len(archivos)

def exportar_consolidad_excel(output_path: str):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 
                d.bloque AS "BLOCK",
                d.numero_depto AS "DEPTO",
                d.rol_sii AS "ROL SII",
                d.avaluo_fiscal AS "AVALÚO FISCAL",
                d.fojas AS "FOJAS",
                d.numero_inscripcion AS "N° INSCRIPCIÓN",
                d.ano_inscripcion AS "AÑO",
                i.parentesco AS "PARENTESCO",
                i.nombres AS "NOMBRES",
                i.apellido_paterno AS "AP. PATERNO",
                i.apellido_materno AS "AP. MATERNO",
                i.rut AS "RUT",
                i.asistencia_reuniones AS "ASISTENCIA",
                d.observaciones AS "OBSERVACIONES"
            FROM departamentos d
            LEFT JOIN integrantes i ON d.id = i.departamento_id
            ORDER BY d.bloque ASC, d.numero_depto ASC;
        """)
        rows = cursor.fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consolidado"

        headers = [
            "BLOCK", "DEPTO", "ROL SII", "AVALÚO FISCAL", "FOJAS", 
            "N° INSCRIPCIÓN", "AÑO", "PARENTESCO", "NOMBRES", 
            "AP. PATERNO", "AP. MATERNO", "RUT", "ASISTENCIA", "OBSERVACIONES"
        ]
        ws.append(headers)

        for row in rows:
            ws.append([
                row["BLOCK"], row["DEPTO"], row["ROL SII"], row["AVALÚO FISCAL"],
                row["FOJAS"], row["N° INSCRIPCIÓN"], row["AÑO"], row["PARENTESCO"],
                row["NOMBRES"], row["AP. PATERNO"], row["AP. MATERNO"], row["RUT"],
                row["ASISTENCIA"], row["OBSERVACIONES"]
            ])

        wb.save(output_path)
        return True, f"Reporte exportado exitosamente a: {output_path}"

    except Exception as e:
        return False, f"Error al exportar: {str(e)}"